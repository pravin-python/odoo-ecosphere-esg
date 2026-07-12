"""Render a :class:`ReportResult` to CSV, XLSX, or PDF bytes.

CSV uses the standard library. XLSX needs ``openpyxl`` and PDF needs
``reportlab`` (both in requirements.txt); if a library is missing we raise a
clear ExportError instead of a cryptic ImportError.
"""
import csv
import io

CONTENT_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


class ExportError(Exception):
    pass


def _slug(title: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in title.lower()).strip("_")


def to_csv(result) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([result.title])
    writer.writerow([])
    writer.writerow(result.columns)
    for row in result.rows:
        writer.writerow([row.get(c, "") for c in result.columns])
    if result.summary:
        writer.writerow([])
        writer.writerow(["Summary"])
        for k, v in result.summary.items():
            writer.writerow([k, v])
    return buf.getvalue().encode("utf-8")


def to_xlsx(result) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ExportError("openpyxl is required for XLSX export.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append([result.title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(result.columns)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for row in result.rows:
        ws.append([row.get(c, "") for c in result.columns])

    if result.summary:
        ws.append([])
        ws.append(["Summary"])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        for k, v in result.summary.items():
            ws.append([k, v])

    for i, col in enumerate(result.columns, start=1):
        if result.rows:
            width = max([len(str(col))] + [len(str(r.get(col, ""))) for r in result.rows])
        else:
            width = len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(result) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover
        raise ExportError("reportlab is required for PDF export.") from exc

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(result.title, styles["Title"]), Spacer(1, 12)]

    data = [result.columns] + [
        [str(row.get(c, "")) for c in result.columns] for row in result.rows
    ]
    if len(data) == 1:
        data.append(["No data" for _ in result.columns])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f8e9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    if result.summary:
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("Summary", styles["Heading2"]))
        for k, v in result.summary.items():
            elements.append(Paragraph(f"<b>{k}:</b> {v}", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


_EXPORTERS = {"csv": to_csv, "xlsx": to_xlsx, "pdf": to_pdf}


def export_report(result, fmt: str):
    """Return ``(bytes, content_type, filename)`` for the requested format."""
    fmt = fmt.lower()
    exporter = _EXPORTERS.get(fmt)
    if exporter is None:
        raise ExportError(f"Unsupported export format: {fmt!r}")
    payload = exporter(result)
    filename = f"{_slug(result.title)}.{fmt}"
    return payload, CONTENT_TYPES[fmt], filename
